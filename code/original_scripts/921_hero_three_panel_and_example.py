#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.stats import beta

REPO_ROOT = Path(__file__).resolve().parents[2]
NI_DIR = REPO_ROOT / 'outputs/reports/ni_gap'
FORCED_DIR = REPO_ROOT / 'outputs/reports/forced_choice'
FIG_DIR = REPO_ROOT / 'outputs/reports/figures'
HUMAN_DIR = REPO_ROOT / 'outputs/reports/human_eval'
HUMAN_INPUT_DIR = HUMAN_DIR / 'annotator_inputs'
MATCHED_DIR = REPO_ROOT / 'outputs/reports/matched_control'
MATCHED_SCORES_DIR = REPO_ROOT / 'outputs/matched_control/scores'
PER_PROBLEM_DIR = REPO_ROOT / 'outputs/reports/aggregation_sweep/per_problem'

PRMS = [
    'skywork_prm',
    'internlm2_7b_reward',
    'qwen_prm',
    'math_shepherd',
    'skywork_prm_1_5b',
    'rlhflow_prm',
    'internlm2_1_8b_reward',
]

DISPLAY_PRM = {
    'skywork_prm': 'Skywork-7B',
    'internlm2_7b_reward': 'InternLM2-7B',
    'qwen_prm': 'Qwen',
    'math_shepherd': 'Math-Shepherd',
    'skywork_prm_1_5b': 'Skywork-1.5B',
    'rlhflow_prm': 'RLHFlow',
    'internlm2_1_8b_reward': 'InternLM2-1.8B',
}
DISPLAY_BENCH = {'gsm8k': 'GSM8K', 'math500': 'MATH-500'}

ANNOTATION_FILES = {
    'skywork_prm': [
        'completed_annotation_sheet_skywork_prm_gsm8k.xlsx',
        'annotation_sheet_skywork_prm_gsm8k_completed.xlsx',
    ],
    'internlm2_7b_reward': [
        'completed_annotation_sheet_internlm2_7b_reward_gsm8k.xlsx',
        'annotation_sheet_internlm2_7b_reward_gsm8k_completed.xlsx',
    ],
}

FORCED_COLORS = {
    'skywork': '#2A6F97',
    'skywork_pooled': '#114B5F',
    'internlm': '#C97C1A',
    'overall': '#444444',
}


@dataclass
class HumanCardSummary:
    card_id: str
    problem_id: str
    prm: str
    product_steps: int
    gm_steps: int
    product_dim_means: dict[str, float]
    gm_dim_means: dict[str, float]
    product_source_candidate_idx: int
    gm_source_candidate_idx: int


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open('r', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open('r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def clopper_pearson(k: int, n: int, alpha: float = 0.05) -> tuple[float, float]:
    if n == 0:
        return float('nan'), float('nan')
    lo = 0.0 if k == 0 else float(beta.ppf(alpha / 2, k, n - k + 1))
    hi = 1.0 if k == n else float(beta.ppf(1 - alpha / 2, k + 1, n - k))
    return lo, hi


def human_label(prm: str, bench: str) -> str:
    return f"{DISPLAY_PRM[prm]} / {DISPLAY_BENCH[bench]}"


def make_three_panel_figure() -> tuple[Path, Path, str]:
    ni_rows = [r for r in load_csv(NI_DIR / 'ni_gap_7prm_product.csv') if r['aggregation'] == 'product']
    matched_rows = load_csv(NI_DIR / 'score_matched_ni.csv')
    contrast_rows = load_csv(NI_DIR / 'decomposition_contrasts.csv')
    fc_rows = load_csv(FORCED_DIR / 'forced_choice_results.csv')

    matched_map = {(r['prm'], r['benchmark']): float(r['delta_ni_matched']) for r in matched_rows}
    contrast_map = {(r['prm'], r['benchmark']): r for r in contrast_rows}

    forest_rows: list[dict[str, Any]] = []
    for r in ni_rows:
        key = (r['prm'], r['benchmark'])
        c = contrast_map[key]
        forest_rows.append({
            'prm': r['prm'],
            'benchmark': r['benchmark'],
            'label': human_label(r['prm'], r['benchmark']),
            'delta_ni': float(r['delta_ni']),
            'ni_ci_lo': float(r['ci_lo']),
            'ni_ci_hi': float(r['ci_hi']),
            'delta_ni_matched': matched_map[key],
            'ctx_contrast': float(c['ctx_contrast']),
            'ctx_ci_lo': float(c['ctx_ci_lo']),
            'ctx_ci_hi': float(c['ctx_ci_hi']),
            'mech_contrast': float(c['mech_contrast']),
            'mech_ci_lo': float(c['mech_ci_lo']),
            'mech_ci_hi': float(c['mech_ci_hi']),
        })

    forest_rows.sort(key=lambda x: x['delta_ni'])

    fc_map = {(r['prm'], r['benchmark']): r for r in fc_rows if r['prm'] != 'pooled'}
    sky_n = int(fc_map[('skywork_prm', 'gsm8k')]['n']) + int(fc_map[('skywork_prm', 'math500')]['n'])
    sky_gm = int(fc_map[('skywork_prm', 'gsm8k')]['gm_preferred']) + int(fc_map[('skywork_prm', 'math500')]['gm_preferred'])
    sky_lo, sky_hi = clopper_pearson(sky_gm, sky_n)
    pooled_row = next(r for r in fc_rows if r['prm'] == 'pooled')

    bar_rows = [
        ('Skywork-7B / GSM8K', float(fc_map[('skywork_prm', 'gsm8k')]['gm_rate']), float(fc_map[('skywork_prm', 'gsm8k')]['ci_lo']), float(fc_map[('skywork_prm', 'gsm8k')]['ci_hi']), FORCED_COLORS['skywork']),
        ('Skywork-7B / MATH-500', float(fc_map[('skywork_prm', 'math500')]['gm_rate']), float(fc_map[('skywork_prm', 'math500')]['ci_lo']), float(fc_map[('skywork_prm', 'math500')]['ci_hi']), FORCED_COLORS['skywork']),
        ('Skywork pooled', sky_gm / sky_n, sky_lo, sky_hi, FORCED_COLORS['skywork_pooled']),
        ('InternLM2-7B / GSM8K', float(fc_map[('internlm2_7b_reward', 'gsm8k')]['gm_rate']), float(fc_map[('internlm2_7b_reward', 'gsm8k')]['ci_lo']), float(fc_map[('internlm2_7b_reward', 'gsm8k')]['ci_hi']), FORCED_COLORS['internlm']),
        ('InternLM2-7B / MATH-500', float(fc_map[('internlm2_7b_reward', 'math500')]['gm_rate']), float(fc_map[('internlm2_7b_reward', 'math500')]['ci_lo']), float(fc_map[('internlm2_7b_reward', 'math500')]['ci_hi']), FORCED_COLORS['internlm']),
        ('Overall pooled', float(pooled_row['gm_rate']), float(pooled_row['ci_lo']), float(pooled_row['ci_hi']), FORCED_COLORS['overall']),
    ]

    plt.rcParams.update({
        'font.size': 9.5,
        'axes.spines.top': False,
        'axes.spines.right': False,
        'pdf.fonttype': 42,
        'ps.fonttype': 42,
    })
    fig, (ax1, ax2, ax3) = plt.subplots(
        1,
        3,
        figsize=(18.0, 8.8),
        gridspec_kw={'width_ratios': [1.25, 1.25, 0.95]},
    )

    y = np.arange(len(forest_rows))

    ni_x = np.array([r['delta_ni'] for r in forest_rows])
    ni_lo = np.array([r['delta_ni'] - r['ni_ci_lo'] for r in forest_rows])
    ni_hi = np.array([r['ni_ci_hi'] - r['delta_ni'] for r in forest_rows])
    matched_x = np.array([r['delta_ni_matched'] for r in forest_rows])

    ax1.errorbar(
        ni_x,
        y,
        xerr=[ni_lo, ni_hi],
        fmt='o',
        color='#1B1F3B',
        ecolor='#1B1F3B',
        elinewidth=1.5,
        capsize=3,
        markersize=4.8,
        label='Raw $\Delta_{NI}$',
    )
    ax1.scatter(
        matched_x,
        y,
        facecolors='white',
        edgecolors='#D55E00',
        s=36,
        linewidths=1.5,
        zorder=3,
        label='Score-matched $\Delta_{NI}$',
    )
    ax1.axvline(0.0, color='#777777', linestyle='--', linewidth=1.1)
    ax1.text(0.004, len(forest_rows) - 0.35, 'Semantic hypothesis predicts $\Delta > 0$', color='#666666', ha='left', va='bottom')
    ax1.set_yticks(y)
    ax1.set_yticklabels([r['label'] for r in forest_rows])
    ax1.invert_yaxis()
    ax1.set_xlim(-0.55, 0.08)
    ax1.set_xlabel('$\Delta_{NI}$ (necessary APR - inert APR)')
    ax1.set_title('A. Necessity Reversal')
    ax1.grid(axis='x', color='#E6E6E6', linewidth=0.7)
    ax1.legend(loc='lower right', frameon=False)

    ctx_x = np.array([r['ctx_contrast'] for r in forest_rows])
    ctx_lo = np.array([r['ctx_contrast'] - r['ctx_ci_lo'] for r in forest_rows])
    ctx_hi = np.array([r['ctx_ci_hi'] - r['ctx_contrast'] for r in forest_rows])
    mech_x = np.array([r['mech_contrast'] for r in forest_rows])
    mech_lo = np.array([r['mech_contrast'] - r['mech_ci_lo'] for r in forest_rows])
    mech_hi = np.array([r['mech_ci_hi'] - r['mech_contrast'] for r in forest_rows])

    ax2.errorbar(
        ctx_x,
        y - 0.11,
        xerr=[ctx_lo, ctx_hi],
        fmt='o',
        color='#0B6E4F',
        ecolor='#0B6E4F',
        elinewidth=1.5,
        capsize=3,
        markersize=4.8,
        label='Rescoring contrast $\Delta_{ctx}^{nec} - \Delta_{ctx}^{inert}$',
    )
    ax2.errorbar(
        mech_x,
        y + 0.11,
        xerr=[mech_lo, mech_hi],
        fmt='o',
        mfc='white',
        mec='#B56576',
        color='#B56576',
        ecolor='#B56576',
        elinewidth=1.3,
        capsize=3,
        markersize=4.6,
        label='Mechanical contrast $\Delta_{mech}^{nec} - \Delta_{mech}^{inert}$',
    )
    ax2.axvline(0.0, color='#777777', linestyle='--', linewidth=1.1)
    ax2.set_yticks(y)
    ax2.set_yticklabels([])
    ax2.invert_yaxis()
    ax2.set_xlim(-1.75, 0.12)
    ax2.set_xlabel('Contrast value')
    ax2.set_title('B. Rescoring vs Mechanical')
    ax2.grid(axis='x', color='#E6E6E6', linewidth=0.7)
    ax2.legend(loc='lower left', frameon=False)

    bar_labels = [r[0] for r in bar_rows]
    bar_vals = np.array([r[1] * 100 for r in bar_rows])
    bar_lo = np.array([(r[1] - r[2]) * 100 for r in bar_rows])
    bar_hi = np.array([(r[3] - r[1]) * 100 for r in bar_rows])
    bar_colors = [r[4] for r in bar_rows]
    y3 = np.arange(len(bar_rows))
    ax3.barh(y3, bar_vals, color=bar_colors, alpha=0.92)
    ax3.errorbar(bar_vals, y3, xerr=[bar_lo, bar_hi], fmt='none', ecolor='black', elinewidth=1.1, capsize=3)
    ax3.axvline(50.0, color='#777777', linestyle='--', linewidth=1.1)
    ax3.text(50.5, len(bar_rows) - 0.45, 'Chance', color='#666666', ha='left', va='bottom')
    ax3.set_yticks(y3)
    ax3.set_yticklabels(bar_labels)
    ax3.invert_yaxis()
    ax3.set_xlim(45, 100)
    ax3.set_xlabel('% GM preferred')
    ax3.set_title('C. Human Preference')
    ax3.grid(axis='x', color='#E6E6E6', linewidth=0.7)

    fig.tight_layout(w_pad=2.8)
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    pdf_path = FIG_DIR / 'hero_three_panel.pdf'
    png_path = FIG_DIR / 'hero_three_panel.png'
    fig.savefig(pdf_path, bbox_inches='tight')
    fig.savefig(png_path, dpi=220, bbox_inches='tight')
    plt.close(fig)

    left_count = sum(1 for r in forest_rows if r['delta_ni'] < 0)
    ctx_sig = sum(1 for r in forest_rows if r['ctx_ci_hi'] < 0)
    mech_cross = sum(1 for r in forest_rows if r['mech_ci_lo'] <= 0 <= r['mech_ci_hi'])
    right_min = min(v for _, v, *_ in bar_rows) * 100
    right_max = max(v for _, v, *_ in bar_rows) * 100
    summary = (
        f'Left panel: {left_count}/14 raw ΔNI point estimates are below zero.\n'
        f'Center panel: rescoring contrast CIs exclude zero in {ctx_sig}/14; mechanical contrast CIs include zero in {mech_cross}/14.\n'
        f'Right panel: GM preference ranges from {right_min:.1f}% to {right_max:.1f}%.\n'
    )
    return pdf_path, png_path, summary


def _parse_human_card(prm: str, card_id: str) -> HumanCardSummary:
    card_rows = load_jsonl(HUMAN_DIR / f'cards_{prm}_gsm8k.jsonl')
    deblind_rows = load_jsonl(HUMAN_DIR / f'deblind_{prm}_gsm8k.jsonl')
    card = next(r for r in card_rows if r['card_id'] == card_id)
    deblind = next(r for r in deblind_rows if r['card_id'] == card_id)

    product_trace_key = next(k for k in ['trace_1', 'trace_2', 'trace_3'] if deblind[f'{k}_source'] == 'product_top1')
    gm_trace_key = next(k for k in ['trace_1', 'trace_2', 'trace_3'] if deblind[f'{k}_source'] == 'gm_top1')

    product_dims = {'completeness': [], 'auditability': [], 'coherence': [], 'supervision': []}
    gm_dims = {'completeness': [], 'auditability': [], 'coherence': [], 'supervision': []}

    for fname in ANNOTATION_FILES[prm]:
        wb = load_workbook(HUMAN_INPUT_DIR / fname, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        row = next(r for r in rows if r[0] == card_id)
        if product_trace_key == 'trace_1':
            p_off = 9
        elif product_trace_key == 'trace_2':
            p_off = 13
        else:
            p_off = 17
        if gm_trace_key == 'trace_1':
            g_off = 9
        elif gm_trace_key == 'trace_2':
            g_off = 13
        else:
            g_off = 17
        product_dims['completeness'].append(float(row[p_off]))
        product_dims['auditability'].append(float(row[p_off + 1]))
        product_dims['coherence'].append(float(row[p_off + 2]))
        product_dims['supervision'].append(float(row[p_off + 3]))
        gm_dims['completeness'].append(float(row[g_off]))
        gm_dims['auditability'].append(float(row[g_off + 1]))
        gm_dims['coherence'].append(float(row[g_off + 2]))
        gm_dims['supervision'].append(float(row[g_off + 3]))

    product_means = {k: float(np.mean(v)) for k, v in product_dims.items()}
    gm_means = {k: float(np.mean(v)) for k, v in gm_dims.items()}

    return HumanCardSummary(
        card_id=card_id,
        problem_id=card['problem_id'],
        prm=prm,
        product_steps=int(card[product_trace_key]['num_steps']),
        gm_steps=int(card[gm_trace_key]['num_steps']),
        product_dim_means=product_means,
        gm_dim_means=gm_means,
        product_source_candidate_idx=int(deblind['product_source_candidate_idx']),
        gm_source_candidate_idx=int(deblind['gm_source_candidate_idx']),
    )


def qualitative_example_summary() -> str:
    nec_rows = load_jsonl(MATCHED_DIR / 'gsm8k_necessary_matched.jsonl')
    inert_rows = load_jsonl(MATCHED_DIR / 'gsm8k_inert_matched.jsonl')
    nec_map = {(r['problem_id'], int(r['candidate_index'])): r for r in nec_rows}
    inert_map = {(r['problem_id'], int(r['candidate_index'])): r for r in inert_rows}
    common_keys = sorted(set(nec_map) & set(inert_map))

    selection_maps = {}
    for prm in PRMS:
        rows = load_jsonl(PER_PROBLEM_DIR / f'gsm8k_{prm}_flip_gain_6rule_per_problem.jsonl')
        selection_maps[prm] = {r['problem_id']: r for r in rows}

    score_maps: dict[tuple[str, str], dict[tuple[str, int], dict[str, Any]]] = {}
    for prm in PRMS:
        score_maps[(prm, 'necessary')] = {(r['problem_id'], int(r['candidate_index'])): r for r in load_jsonl(MATCHED_SCORES_DIR / f'{prm}_gsm8k_necessary.jsonl')}
        score_maps[(prm, 'inert')] = {(r['problem_id'], int(r['candidate_index'])): r for r in load_jsonl(MATCHED_SCORES_DIR / f'{prm}_gsm8k_inert.jsonl')}

    human_exact = set()
    human_problem_ids = set()
    for prm in ['skywork_prm', 'internlm2_7b_reward', 'qwen_prm']:
        for d in load_jsonl(HUMAN_DIR / f'deblind_{prm}_gsm8k.jsonl'):
            if d.get('product_flip_dest') == 'to_deleted' and int(d['product_source_candidate_idx']) == int(d['gm_source_candidate_idx']):
                key = (d['problem_id'], int(d['product_source_candidate_idx']))
                human_exact.add((prm, key))
                human_problem_ids.add(d['problem_id'])

    exact_hits: list[dict[str, Any]] = []
    mechanistic_hits: list[dict[str, Any]] = []
    for prm in PRMS:
        for key in common_keys:
            pid, cidx = key
            s = selection_maps[prm].get(pid)
            if not s:
                continue
            nec_score = score_maps[(prm, 'necessary')].get(key)
            inert_score = score_maps[(prm, 'inert')].get(key)
            if not nec_score or not inert_score:
                continue
            cp = float(nec_score['traj_product_clean'])
            np_ = float(nec_score['traj_product_modified'])
            ip = float(inert_score['traj_product_modified'])
            cg = float(nec_score['traj_gm_clean'])
            ng = float(nec_score['traj_gm_modified'])
            ig = float(inert_score['traj_gm_modified'])
            product_pattern = ip > cp and np_ < cp
            gm_clean_highest = cg >= ig and cg >= ng
            selected_by_both = s.get('product_selected_idx') == cidx and s.get('gm_selected_idx') == cidx
            row = {
                'prm': prm,
                'problem_id': pid,
                'candidate_index': cidx,
                'clean_prod': cp,
                'inert_prod': ip,
                'necessary_prod': np_,
                'clean_gm': cg,
                'inert_gm': ig,
                'necessary_gm': ng,
                'selected_by_both': selected_by_both,
                'product_pattern': product_pattern,
                'gm_clean_highest': gm_clean_highest,
                'human_exact': (prm, key) in human_exact,
                'human_problem_overlap': pid in human_problem_ids,
                'n_steps_clean': int(nec_map[key]['n_steps_clean']),
                'inert_deleted_step': int(inert_map[key]['deleted_step_index_1based']),
                'necessary_deleted_step': int(nec_map[key]['deleted_step_index_1based']),
            }
            if product_pattern and gm_clean_highest:
                mechanistic_hits.append(row)
            if product_pattern and gm_clean_highest and selected_by_both and (prm, key) in human_exact:
                exact_hits.append(row)

    lines = []
    lines.append('Qualitative worked-example search')
    lines.append('')
    lines.append('Exact target conditions:')
    lines.append('1. same clean candidate has both necessary and inert deletions in matched_control')
    lines.append('2. on the clean pool, product and GM both select that candidate')
    lines.append('3. product prefers inert-delete over clean but not necessary-delete')
    lines.append('4. the same candidate appears in GSM8K human-eval cards with a deletion-stress judgment')
    lines.append('')
    lines.append(f'Common same-candidate matched-control keys: {len(common_keys)}')
    lines.append(f'Human-eval exact candidate overlaps on deletion-stress cards: {sum(1 for prm_key in human_exact if prm_key[1] in common_keys)}')
    lines.append(f'Exact four-condition hits: {len(exact_hits)}')
    lines.append('')

    if exact_hits:
        lines.append('Exact hit(s):')
        for hit in exact_hits:
            lines.append(json.dumps(hit, ensure_ascii=False))
        return '\n'.join(lines) + '\n'

    lines.append('No exact worked example exists in the current logged artifacts.')
    lines.append('There are two independent blockers:')
    lines.append('- none of the 17 GSM8K common matched-control candidates are also used as same-candidate deletion-stress human-eval cards;')
    lines.append('- after additionally requiring clean-time selection by both product and GM, the hit count remains zero.')
    lines.append('')

    same_problem_hits = [r for r in mechanistic_hits if r['human_problem_overlap']]
    same_problem_hits.sort(key=lambda r: ((r['inert_prod'] - r['clean_prod']) + (r['clean_prod'] - r['necessary_prod'])), reverse=True)
    if same_problem_hits:
        pick = same_problem_hits[0]
        pid = pick['problem_id']
        cidx = pick['candidate_index']
        nr = nec_map[(pid, cidx)]
        ir = inert_map[(pid, cidx)]
        human_card = _parse_human_card('skywork_prm', 'gsm8k_skywork_prm_006') if pid == 'gsm8k_112' else None

        lines.append('Nearest same-problem exemplar (candidate mismatch, so use only as an appendix anecdote):')
        lines.append(f"- Problem: {pid}")
        lines.append(f"- Question: {nr['original_trace']['problem']['question']}")
        lines.append(f"- Matched-control candidate index: {cidx}")
        lines.append(f"- Clean steps: {pick['n_steps_clean']}; inert deletion removes step {pick['inert_deleted_step']}; necessary deletion removes step {pick['necessary_deleted_step']}")
        lines.append(
            f"- {DISPLAY_PRM[pick['prm']]} scores: product clean={pick['clean_prod']:.4f}, inert={pick['inert_prod']:.4f}, necessary={pick['necessary_prod']:.4f}; "
            f"GM clean={pick['clean_gm']:.4f}, inert={pick['inert_gm']:.4f}, necessary={pick['necessary_gm']:.4f}"
        )
        if human_card is not None:
            lines.append('- Same-problem human card exists, but on a different candidate (candidate 13).')
            lines.append(
                f"- Human card {human_card.card_id}: product trace {human_card.product_steps} steps vs GM trace {human_card.gm_steps} steps; "
                f"mean completeness {human_card.product_dim_means['completeness']:.1f} vs {human_card.gm_dim_means['completeness']:.1f}, "
                f"auditability {human_card.product_dim_means['auditability']:.1f} vs {human_card.gm_dim_means['auditability']:.1f}, "
                f"coherence {human_card.product_dim_means['coherence']:.1f} vs {human_card.gm_dim_means['coherence']:.1f}, "
                f"supervision {human_card.product_dim_means['supervision']:.1f} vs {human_card.gm_dim_means['supervision']:.1f}."
            )
        lines.append('')
        lines.append('Recommendation: use the three-panel figure as the main exhibit, and only include a worked example if you explicitly label it as a near-miss / same-problem anecdote rather than an exact single-card witness.')

    return '\n'.join(lines) + '\n'


def main() -> int:
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    pdf_path, png_path, fig_summary = make_three_panel_figure()
    (FIG_DIR / 'hero_three_panel_summary.txt').write_text(fig_summary, encoding='utf-8')
    qual_summary = qualitative_example_summary()
    (FIG_DIR / 'qualitative_example_search_summary.txt').write_text(qual_summary, encoding='utf-8')
    print(fig_summary.strip())
    print('---')
    print(qual_summary.strip())
    print('---')
    print(pdf_path)
    print(png_path)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
