#!/usr/bin/env python3
from __future__ import annotations

import csv
import filecmp
import json
import math
import random
import shutil
import sys
import tempfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RELEASE_ROOT = Path(__file__).resolve().parent
ROOT = RELEASE_ROOT / "human_cards/no_preference_reannotation"
INCOMING = ROOT / "completed"
PUBLISHED = ROOT / "analysis"
OUT = Path(tempfile.mkdtemp(prefix="strictprmbench_no_pref_repro_"))
SEED = 42
BOOTSTRAPS = 10000
LABELS = ["gm", "product", "no_preference"]


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows and fieldnames is None:
        raise ValueError(f"No rows and no fieldnames for {path}")
    if fieldnames is None:
        fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def normalize_choice(raw: Any) -> str:
    if raw is None:
        return ""
    text = str(raw).strip()
    low = text.lower().replace("_", " ").replace("-", " ")
    low = " ".join(low.split())
    if low in {"a", "trace a", "choice a"}:
        return "A"
    if low in {"b", "trace b", "choice b"}:
        return "B"
    if low in {"no preference", "nopreference", "no pref", "tie", "equal"}:
        return "No preference"
    return text


def read_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Annotations" not in wb.sheetnames:
        raise RuntimeError(f"{path} missing Annotations sheet; sheets={wb.sheetnames}")
    ws = wb["Annotations"]
    headers = [cell.value for cell in ws[1]]
    required = ["row_num", "card_id", "problem_statement", "trace_A", "trace_B", "preference", "notes"]
    if headers[: len(required)] != required:
        raise RuntimeError(f"{path} unexpected headers: {headers}")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def load_manifest() -> dict[str, list[dict[str, str]]]:
    out: dict[str, list[dict[str, str]]] = defaultdict(list)
    with (ROOT / "per_annotator_manifest.csv").open("r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[str(row["annotator"])].append(row)
    for rows in out.values():
        rows.sort(key=lambda r: int(r["row_num"]))
    return dict(out)


def map_workbooks(paths: list[Path], manifest: dict[str, list[dict[str, str]]]) -> tuple[dict[Path, str], list[str]]:
    audit_lines = []
    candidates: list[tuple[int, Path, str, int, int]] = []
    for path in paths:
        rows = read_workbook(path)
        order = [str(r["card_id"]) for r in rows]
        set_order = set(order)
        for ann, man_rows in manifest.items():
            man_order = [r["card_id"] for r in man_rows]
            exact_positions = sum(1 for a, b in zip(order, man_order) if a == b)
            set_overlap = len(set_order & set(man_order))
            candidates.append((exact_positions, path, ann, set_overlap, len(rows)))
            audit_lines.append(
                f"mapping_candidate file={path.name} annotator={ann} exact_positions={exact_positions} set_overlap={set_overlap} rows={len(rows)}"
            )
    candidates.sort(reverse=True, key=lambda x: (x[0], x[3]))
    used_paths: set[Path] = set()
    used_anns: set[str] = set()
    mapping: dict[Path, str] = {}
    for exact, path, ann, overlap, n_rows in candidates:
        if path in used_paths or ann in used_anns:
            continue
        if exact < 200:
            raise RuntimeError(f"Cannot confidently map {path}: best exact match {exact}/200")
        mapping[path] = ann
        used_paths.add(path)
        used_anns.add(ann)
        audit_lines.append(f"mapped file={path.name} -> annotator={ann} exact_positions={exact} set_overlap={overlap}")
    if len(mapping) != 3:
        raise RuntimeError(f"Expected 3 mapped workbooks, got {len(mapping)}")
    return mapping, audit_lines


def deblind_choice(choice: str, assignment: str) -> str:
    if choice == "No preference":
        return "no_preference"
    if choice not in {"A", "B"}:
        return "invalid"
    if assignment == "product->A,gm->B":
        return "product" if choice == "A" else "gm"
    if assignment == "gm->A,product->B":
        return "gm" if choice == "A" else "product"
    raise RuntimeError(f"Unknown assignment {assignment}")


def bootstrap_rate(bits: list[int], seed_key: str) -> tuple[float, float, float]:
    if not bits:
        return float("nan"), float("nan"), float("nan")
    rng = random.Random(f"{SEED}:{seed_key}")
    n = len(bits)
    point = sum(bits) / n
    boots = []
    for _ in range(BOOTSTRAPS):
        boots.append(sum(bits[rng.randrange(n)] for _ in range(n)) / n)
    boots.sort()
    return point, boots[int(0.025 * BOOTSTRAPS)], boots[int(0.975 * BOOTSTRAPS)]


def binom_two_sided(k: int, n: int, p: float = 0.5) -> float:
    if n == 0:
        return float("nan")
    try:
        from scipy.stats import binomtest

        return float(binomtest(k, n, p=p, alternative="two-sided").pvalue)
    except Exception:
        probs = [math.comb(n, i) * (p**i) * ((1 - p) ** (n - i)) for i in range(n + 1)]
        obs = probs[k]
        return float(sum(prob for prob in probs if prob <= obs + 1e-15))


def summarize_cards(cards: list[dict[str, Any]], group_name: str) -> dict[str, Any]:
    n = len(cards)
    counts = Counter(c["majority_label"] for c in cards)
    resolved = [c for c in cards if c["majority_label"] in {"gm", "product"}]
    resolved_gm = sum(1 for c in resolved if c["majority_label"] == "gm")
    no_pref_bits = [1 if c["majority_label"] == "no_preference" else 0 for c in cards]
    gm_raw_bits = [1 if c["majority_label"] == "gm" else 0 for c in cards]
    product_raw_bits = [1 if c["majority_label"] == "product" else 0 for c in cards]
    no_pref_point, no_pref_lo, no_pref_hi = bootstrap_rate(no_pref_bits, f"{group_name}:no_pref")
    gm_raw_point, gm_raw_lo, gm_raw_hi = bootstrap_rate(gm_raw_bits, f"{group_name}:gm_raw")
    product_raw_point, product_raw_lo, product_raw_hi = bootstrap_rate(product_raw_bits, f"{group_name}:product_raw")
    if resolved:
        resolved_bits = [1 if c["majority_label"] == "gm" else 0 for c in resolved]
        resolved_point, resolved_lo, resolved_hi = bootstrap_rate(resolved_bits, f"{group_name}:resolved")
    else:
        resolved_point = resolved_lo = resolved_hi = float("nan")
    return {
        "group": group_name,
        "n_cards": n,
        "gm_majority": counts["gm"],
        "product_majority": counts["product"],
        "no_preference_majority": counts["no_preference"],
        "split": counts["split"],
        "resolved_n": len(resolved),
        "resolved_gm": resolved_gm,
        "resolved_product": len(resolved) - resolved_gm,
        "resolved_gm_rate": resolved_point,
        "resolved_gm_ci_low": resolved_lo,
        "resolved_gm_ci_high": resolved_hi,
        "resolved_binom_p_vs_50": binom_two_sided(resolved_gm, len(resolved)),
        "no_preference_rate": no_pref_point,
        "no_preference_ci_low": no_pref_lo,
        "no_preference_ci_high": no_pref_hi,
        "gm_raw_rate": gm_raw_point,
        "gm_raw_ci_low": gm_raw_lo,
        "gm_raw_ci_high": gm_raw_hi,
        "product_raw_rate": product_raw_point,
        "product_raw_ci_low": product_raw_lo,
        "product_raw_ci_high": product_raw_hi,
    }


def fleiss_kappa(card_labels: dict[str, list[str]]) -> float:
    n_cards = len(card_labels)
    n_raters = 3
    p_j = {label: 0 for label in LABELS}
    p_i = []
    for labels in card_labels.values():
        counts = Counter(labels)
        for label in LABELS:
            p_j[label] += counts[label]
        p_i.append((sum(counts[label] ** 2 for label in LABELS) - n_raters) / (n_raters * (n_raters - 1)))
    p_bar = sum(p_i) / n_cards
    p_e = sum((p_j[label] / (n_cards * n_raters)) ** 2 for label in LABELS)
    if abs(1 - p_e) < 1e-12:
        return float("nan")
    return (p_bar - p_e) / (1 - p_e)


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    deblind = {r["card_id"]: r for r in load_jsonl(ROOT / "deblind_key.jsonl")}
    manifest = load_manifest()
    incoming = sorted(INCOMING.glob("*.xlsx"))
    if len(incoming) != 3:
        raise RuntimeError(f"Expected 3 xlsx files in {INCOMING}, found {len(incoming)}")
    mapping, mapping_audit = map_workbooks(incoming, manifest)

    manifest_lookup = {
        (ann, row["card_id"]): row for ann, rows in manifest.items() for row in rows
    }
    per_response: list[dict[str, Any]] = []
    audit_lines = list(mapping_audit)
    invalids = []
    missing = []
    duplicate_checks = []

    for path, ann in sorted(mapping.items(), key=lambda x: x[1]):
        rows = read_workbook(path)
        card_ids = [str(r["card_id"]) for r in rows]
        duplicate_checks.append(f"annotator={ann} file={path.name} rows={len(rows)} unique_card_ids={len(set(card_ids))}")
        for row in rows:
            card_id = str(row["card_id"])
            raw_pref = row.get("preference")
            pref = normalize_choice(raw_pref)
            if pref == "":
                missing.append((ann, card_id, row.get("row_num")))
            if pref not in {"A", "B", "No preference"}:
                invalids.append((ann, card_id, row.get("row_num"), raw_pref))
            man = manifest_lookup[(ann, card_id)]
            label = deblind_choice(pref, man["assignment"]) if pref in {"A", "B", "No preference"} else "invalid"
            meta = deblind[card_id]
            per_response.append(
                {
                    "annotator": ann,
                    "source_file": path.name,
                    "row_num": int(row["row_num"]),
                    "card_id": card_id,
                    "source_card_id": meta["source_card_id"],
                    "prm": meta["prm"],
                    "benchmark": meta["benchmark"],
                    "problem_id": meta["problem_id"],
                    "assignment": man["assignment"],
                    "preference_raw": "" if raw_pref is None else str(raw_pref),
                    "preference_norm": pref,
                    "deblind_label": label,
                    "notes": "" if row.get("notes") is None else str(row.get("notes")),
                }
            )

    if missing or invalids:
        audit_lines.append(f"missing_preferences={missing[:20]}")
        audit_lines.append(f"invalid_preferences={invalids[:20]}")
        raise RuntimeError(f"Found missing={len(missing)} invalid={len(invalids)} preferences")

    card_to_labels: dict[str, list[str]] = defaultdict(list)
    card_to_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in per_response:
        card_to_labels[row["card_id"]].append(row["deblind_label"])
        card_to_rows[row["card_id"]].append(row)

    card_rows: list[dict[str, Any]] = []
    for card_id, labels in sorted(card_to_labels.items()):
        counts = Counter(labels)
        if len(labels) != 3:
            majority = "incomplete"
        else:
            top = counts.most_common()
            majority = top[0][0] if top[0][1] >= 2 else "split"
        meta = deblind[card_id]
        card_rows.append(
            {
                "card_id": card_id,
                "source_card_id": meta["source_card_id"],
                "prm": meta["prm"],
                "benchmark": meta["benchmark"],
                "problem_id": meta["problem_id"],
                "ann_labels": "|".join(labels),
                "gm_votes": counts["gm"],
                "product_votes": counts["product"],
                "no_preference_votes": counts["no_preference"],
                "majority_label": majority,
            }
        )

    summary_rows: list[dict[str, Any]] = []
    summary_rows.append(summarize_cards(card_rows, "overall"))
    for prm in sorted({r["prm"] for r in card_rows}):
        summary_rows.append(summarize_cards([r for r in card_rows if r["prm"] == prm], f"prm={prm}"))
    for bench in sorted({r["benchmark"] for r in card_rows}):
        summary_rows.append(summarize_cards([r for r in card_rows if r["benchmark"] == bench], f"benchmark={bench}"))
    for prm in sorted({r["prm"] for r in card_rows}):
        for bench in sorted({r["benchmark"] for r in card_rows}):
            subset = [r for r in card_rows if r["prm"] == prm and r["benchmark"] == bench]
            if subset:
                summary_rows.append(summarize_cards(subset, f"prm={prm};benchmark={bench}"))

    pair_rows = []
    ann_ids = sorted({r["annotator"] for r in per_response})
    label_by_ann = {
        ann: {r["card_id"]: r["deblind_label"] for r in per_response if r["annotator"] == ann}
        for ann in ann_ids
    }
    for i, a in enumerate(ann_ids):
        for b in ann_ids[i + 1 :]:
            common = sorted(set(label_by_ann[a]) & set(label_by_ann[b]))
            same = sum(1 for cid in common if label_by_ann[a][cid] == label_by_ann[b][cid])
            pair_rows.append(
                {
                    "annotator_a": a,
                    "annotator_b": b,
                    "n_common": len(common),
                    "same_deblind_label": same,
                    "agreement_rate": same / len(common) if common else float("nan"),
                }
            )

    kappa = fleiss_kappa(card_to_labels)

    write_csv(OUT / "per_response.csv", per_response)
    write_csv(OUT / "card_majority.csv", card_rows)
    write_csv(OUT / "majority_summary.csv", summary_rows)
    write_csv(OUT / "pairwise_agreement.csv", pair_rows)

    overall = summary_rows[0]
    audit_lines.extend(duplicate_checks)
    audit_lines.append(f"mapped_files={{{', '.join(f'{p.name}:ann{a}' for p, a in sorted(mapping.items(), key=lambda x: x[1]))}}}")
    audit_lines.append(f"per_response_rows={len(per_response)}")
    audit_lines.append(f"card_rows={len(card_rows)}")
    audit_lines.append(f"missing_preferences={len(missing)}")
    audit_lines.append(f"invalid_preferences={len(invalids)}")
    audit_lines.append(f"fleiss_kappa={kappa:.6f}")
    (OUT / "input_audit.txt").write_text("\n".join(audit_lines) + "\n", encoding="utf-8")

    lines = [
        "No-preference forced-choice reannotation analysis",
        "",
        "Input audit:",
        f"- Workbooks mapped: {', '.join(f'{p.name}->annotator_{a}' for p, a in sorted(mapping.items(), key=lambda x: x[1]))}",
        f"- Responses: {len(per_response)} / 600",
        f"- Cards: {len(card_rows)} / 200",
        f"- Missing preferences: {len(missing)}",
        f"- Invalid preferences: {len(invalids)}",
        f"- Fleiss' kappa: {kappa:.3f}",
        "",
        "Overall majority outcome:",
        f"- GM majority: {overall['gm_majority']}/200 = {100*overall['gm_majority']/200:.1f}%",
        f"- Product majority: {overall['product_majority']}/200 = {100*overall['product_majority']/200:.1f}%",
        f"- No-preference majority: {overall['no_preference_majority']}/200 = {100*overall['no_preference_majority']/200:.1f}%",
        f"- Split: {overall['split']}/200 = {100*overall['split']/200:.1f}%",
        "",
        "Resolved-card GM preference:",
        f"- {overall['resolved_gm']}/{overall['resolved_n']} = {100*overall['resolved_gm_rate']:.1f}%",
        f"- 95% bootstrap CI: [{100*overall['resolved_gm_ci_low']:.1f}, {100*overall['resolved_gm_ci_high']:.1f}]",
        f"- binomial p vs 50%: {overall['resolved_binom_p_vs_50']:.6g}",
        "",
        "No-preference rate:",
        f"- {100*overall['no_preference_rate']:.1f}%",
        f"- 95% bootstrap CI: [{100*overall['no_preference_ci_low']:.1f}, {100*overall['no_preference_ci_high']:.1f}]",
        "",
        "Pairwise agreement:",
    ]
    for row in pair_rows:
        lines.append(
            f"- annotator_{row['annotator_a']} vs annotator_{row['annotator_b']}: "
            f"{row['same_deblind_label']}/{row['n_common']} = {100*row['agreement_rate']:.1f}%"
        )
    lines.extend(["", "Files:", "- per_response.csv", "- card_majority.csv", "- majority_summary.csv", "- pairwise_agreement.csv", "- input_audit.txt"])
    (OUT / "summary.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("\n".join(lines))
    return 0



def check_reproduced_outputs() -> None:
    expected = [
        "per_response.csv",
        "card_majority.csv",
        "majority_summary.csv",
        "pairwise_agreement.csv",
        "input_audit.txt",
        "summary.txt",
    ]
    mismatches = []
    for name in expected:
        generated = OUT / name
        published = PUBLISHED / name
        if not published.exists():
            mismatches.append(f"missing published file: {published}")
        elif not generated.exists():
            mismatches.append(f"missing generated file: {generated}")
        elif not filecmp.cmp(generated, published, shallow=False):
            mismatches.append(f"content mismatch: {name}")
    if mismatches:
        raise RuntimeError("No-preference reannotation reproduction failed: " + "; ".join(mismatches))
    print("PASSED: no-preference reannotation outputs match released analysis files")


if __name__ == "__main__":
    try:
        code = main()
        if "--check" in sys.argv:
            check_reproduced_outputs()
        raise SystemExit(code)
    finally:
        shutil.rmtree(OUT, ignore_errors=True)
